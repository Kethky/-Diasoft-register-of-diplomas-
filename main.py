#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Главный файл системы проверки дипломов
Включает управление ВУЗами, студентами, дипломами, QR-кодами,
временными ссылками, rate limiting, защитой от брутфорса,
а также цифровую подпись дипломов (RSA).
"""
import json
from dotenv import load_dotenv
import os
import sys
import sqlite3
import bcrypt
import time
import qrcode
import base64
import tempfile
import cv2
import pandas as pd
import io
import hashlib
import secrets
from datetime import datetime, timedelta
from functools import wraps, lru_cache
import openpyxl
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.serialization import load_pem_private_key, load_pem_public_key
from cryptography.hazmat.backends import default_backend

load_dotenv()

# ---------- Инициализация приложения ----------
app = Flask(__name__)
app.secret_key = os.urandom(24).hex()
app.config['SECRET_KEY'] = os.urandom(24).hex()
DB_PATH = os.path.join(os.path.dirname(__file__), "database", "diploma_platform.db")
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# ---------- Настройка rate limiter ----------
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["1000 per day", "200 per hour"],
    storage_uri="memory://",
)
limiter.init_app(app)

# ---------- Функции для работы с базой данных ----------
def update_db_schema():
    """Добавляет необходимые поля в таблицу diplomas"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(diplomas)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'active_token' not in columns:
        print("📦 Добавляем поле active_token...")
        cursor.execute("ALTER TABLE diplomas ADD COLUMN active_token TEXT")
        conn.commit()
        print("✅ Поле active_token добавлено")
    if 'hash_combined' not in columns:
        print("📦 Добавляем поле hash_combined...")
        cursor.execute("ALTER TABLE diplomas ADD COLUMN hash_combined TEXT")
        conn.commit()
        print("✅ Поле hash_combined добавлено")
    if 'digital_signature' not in columns:
        print("📦 Добавляем поле digital_signature...")
        cursor.execute("ALTER TABLE diplomas ADD COLUMN digital_signature TEXT")
        conn.commit()
        print("✅ Поле digital_signature добавлено")
    if 'signature_created_at' not in columns:
        print("📦 Добавляем поле signature_created_at...")
        cursor.execute("ALTER TABLE diplomas ADD COLUMN signature_created_at TIMESTAMP")
        conn.commit()
        print("✅ Поле signature_created_at добавлено")
    conn.close()


def log_verification(diploma_id, university_code, diploma_number, verification_type, token, result):
    """Записывает информацию о проверке диплома"""
    ip = request.remote_addr
    user_agent = request.headers.get('User-Agent', 'Unknown')

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO verification_logs (diploma_id, university_code, diploma_number, verification_type, token, ip, user_agent, result)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (diploma_id, university_code, diploma_number, verification_type, token, ip, user_agent, result))
    conn.commit()
    conn.close()

def create_security_tables():
    """Создаёт таблицы для безопасности (блокировка IP, логи)"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS blocked_ips (
            ip TEXT PRIMARY KEY,
            reason TEXT,
            blocked_until TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS security_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ip TEXT,
            endpoint TEXT,
            user_agent TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # НОВАЯ ТАБЛИЦА: логи проверок дипломов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS verification_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            diploma_id INTEGER,
            university_code INTEGER,
            diploma_number TEXT,
            verification_type TEXT,
            token TEXT,
            ip TEXT,
            user_agent TEXT,
            result TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()

@lru_cache(maxsize=1)
def get_university_list():
    """Возвращает список всех вузов"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT university_code, name FROM universities ORDER BY name")
    rows = cursor.fetchall()
    conn.close()
    return [{"code": row[0], "name": row[1]} for row in rows]

def check_university_auth(login: str, password: str) -> dict:
    """Проверяет логин и пароль ВУЗа."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT university_code, name, password_hash FROM universities WHERE login = ?", (login,))
    row = cursor.fetchone()
    conn.close()
    if row:
        university_code, name, stored_hash = row
        if bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8')):
            return {"success": True, "university_code": university_code, "name": name, "message": "Вход выполнен"}
        else:
            return {"success": False, "message": "Неверный пароль"}
    else:
        return {"success": False, "message": "Логин не найден"}

def save_suspicious_report(university_code: int, diploma_number: str, comment: str):
    report = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "university_code": university_code,
        "diploma_number": diploma_number,
        "comment": comment
    }

    log_file = os.path.join(os.path.dirname(__file__), "suspicious_reports.log")
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(json.dumps(report, ensure_ascii=False) + "\n")

def check_student_auth(full_name: str, diploma_number: str) -> bool:
    """Проверяет студента по ФИО и номеру диплома"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM diplomas WHERE full_name = ? AND diploma_number = ?', (full_name, diploma_number))
    row = cursor.fetchone()
    conn.close()
    return row is not None

def get_university_diplomas(university_code: int):
    """Возвращает список дипломов для конкретного ВУЗа"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, full_name, graduation_year, specialty, diploma_number, status
        FROM diplomas WHERE university_code = ? ORDER BY id
    ''', (university_code,))
    rows = cursor.fetchall()
    conn.close()
    return [{"id": r[0], "full_name": r[1], "graduation_year": r[2],
             "specialty": r[3], "diploma_number": r[4], "status": r[5]} for r in rows]

def update_diploma_status(diploma_id: int, new_status: int, university_code: int) -> bool:
    """Обновляет статус диплома, проверяя принадлежность ВУЗу"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE diplomas SET status = ? WHERE id = ? AND university_code = ?',
                   (new_status, diploma_id, university_code))
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return affected > 0

# ---------- Криптографические функции (цифровая подпись) ----------
def ensure_university_keys(university_code: int):
    """Проверяет наличие ключей для ВУЗа, при отсутствии генерирует новые."""
    keys_dir = f"keys/university_{university_code}"
    private_path = f"{keys_dir}/private_key.pem"
    public_path = f"{keys_dir}/public_key.pem"
    if os.path.exists(private_path) and os.path.exists(public_path):
        return True
    os.makedirs(keys_dir, exist_ok=True)
    private_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=2048,
        backend=default_backend()
    )
    with open(private_path, "wb") as f:
        f.write(private_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.NoEncryption()
        ))
    public_key = private_key.public_key()
    with open(public_path, "wb") as f:
        f.write(public_key.public_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PublicFormat.SubjectPublicKeyInfo
        ))
    print(f"✅ Сгенерированы ключи для ВУЗа {university_code}")
    return True

def calculate_diploma_hash(university_code: int, diploma_number: str, full_name: str,
                           graduation_year: int, specialty: str, salt: str = None) -> tuple:
    """Вычисляет хеш диплома. Возвращает (hash_combined, salt)."""
    if salt is None:
        salt = secrets.token_hex(32)
    raw_data = f"{university_code}|{diploma_number}|{full_name}|{graduation_year}|{specialty}|{salt}"
    hash_combined = hashlib.sha256(raw_data.encode()).hexdigest()
    return hash_combined, salt

def sign_diploma(university_code: int, diploma_data: dict) -> str:
    """Подписывает диплом приватным ключом ВУЗа. Возвращает подпись в base64."""
    sign_string = f"{diploma_data['university_code']}|{diploma_data['diploma_number']}|{diploma_data['full_name']}|{diploma_data['graduation_year']}|{diploma_data['specialty']}"
    ensure_university_keys(university_code)
    keys_dir = f"keys/university_{university_code}"
    with open(f"{keys_dir}/private_key.pem", "rb") as f:
        private_key = load_pem_private_key(f.read(), password=None)
    signature = private_key.sign(
        sign_string.encode(),
        padding.PSS(
            mgf=padding.MGF1(hashes.SHA256()),
            salt_length=padding.PSS.MAX_LENGTH
        ),
        hashes.SHA256()
    )
    return base64.b64encode(signature).decode()

def verify_diploma_signature(university_code: int, diploma_data: dict, signature_b64: str) -> bool:
    """Проверяет подпись диплома. Возвращает True/False."""
    sign_string = f"{diploma_data['university_code']}|{diploma_data['diploma_number']}|{diploma_data['full_name']}|{diploma_data['graduation_year']}|{diploma_data['specialty']}"
    keys_dir = f"keys/university_{university_code}"
    public_path = f"{keys_dir}/public_key.pem"
    if not os.path.exists(public_path):
        return False
    with open(public_path, "rb") as f:
        public_key = load_pem_public_key(f.read())
    try:
        signature = base64.b64decode(signature_b64)
        public_key.verify(
            signature,
            sign_string.encode(),
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )
        return True
    except Exception:
        return False

def check_diploma_by_params(university_code: int, diploma_number: str) -> dict:
    """Проверяет диплом по коду вуза и номеру диплома, возвращает также подпись"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT full_name, status, graduation_year, specialty, digital_signature
        FROM diplomas WHERE university_code = ? AND diploma_number = ?
    ''', (university_code, diploma_number))
    row = cursor.fetchone()
    conn.close()
    if row:
        full_name, status, graduation_year, specialty, digital_signature = row
        return {
            "found": True, "valid": status == 1,
            "full_name": full_name, "graduation_year": graduation_year,
            "specialty": specialty, "message": "Подлинный" if status == 1 else "Аннулирован",
            "digital_signature": digital_signature
        }
    else:
        return {"found": False, "valid": False, "message": "Сведений не найдено"}

# ---------- Функции для rate limiting и безопасности ----------
def is_ip_blocked(ip):
    """Проверяет, заблокирован ли IP"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT blocked_until FROM blocked_ips WHERE ip = ? AND blocked_until > datetime("now")', (ip,))
    row = cursor.fetchone()
    conn.close()
    return row is not None

def log_security_event(ip, endpoint, user_agent=None):
    """Логирует подозрительную активность"""
    if user_agent is None:
        user_agent = request.headers.get('User-Agent', 'Unknown')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('INSERT INTO security_logs (ip, endpoint, user_agent) VALUES (?, ?, ?)',
                   (ip, endpoint, user_agent))
    conn.commit()
    conn.close()

def block_ip(ip, reason, minutes=30):
    """Блокирует IP на указанное количество минут"""
    blocked_until = (datetime.now() + timedelta(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('INSERT OR REPLACE INTO blocked_ips (ip, reason, blocked_until) VALUES (?, ?, ?)',
                   (ip, reason, blocked_until))
    conn.commit()
    conn.close()
    print(f"🚫 IP {ip} заблокирован до {blocked_until}. Причина: {reason}")

def check_blocked_ip(f):
    """Декоратор для проверки блокировки IP"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        ip = request.remote_addr
        if is_ip_blocked(ip):
            return jsonify({"error": "IP заблокирован за нарушение правил использования. Обратитесь к администратору."}), 403
        return f(*args, **kwargs)
    return decorated_function

# ---------- Обработчик превышения лимитов ----------
@app.errorhandler(429)
def ratelimit_handler(e):
    ip = request.remote_addr
    log_security_event(ip, request.endpoint)
    return jsonify({"error": "Слишком много запросов", "message": str(e.description), "retry_after": 60}), 429

# ---------- Маршруты (URL) ----------
@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/index.html')
def index_html():
    return redirect(url_for('index'))

@app.route('/dashboard.html')
def dashboard_html():
    return redirect(url_for('dashboard'))

@app.route('/university')
def university():
    if 'role' not in session or session.get('role') != 'university':
        return redirect(url_for('index'))
    university_code = session.get('university_code')
    university_name = session.get('university_name')
    diplomas = get_university_diplomas(university_code)
    return render_template('university.html', university_name=university_name,
                           university_code=university_code, diplomas=diplomas)

@app.route('/student')
def student():
    if 'role' not in session or session.get('role') != 'student':
        return redirect(url_for('index'))
    full_name = session.get('full_name')
    diploma_number = session.get('diploma_number')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT status, graduation_year, specialty, university_code
        FROM diplomas WHERE full_name = ? AND diploma_number = ?
    ''', (full_name, diploma_number))
    row = cursor.fetchone()
    conn.close()
    if not row:
        session.clear()
        return redirect(url_for('index'))
    status, graduation_year, specialty, university_code = row
    return render_template('student.html', full_name=full_name, diploma_number=diploma_number,
                           status=status, graduation_year=graduation_year,
                           specialty=specialty, university_code=university_code)

# ---------- API для ВУЗов ----------
@app.route('/api/universities')
def api_get_universities():
    universities = get_university_list()
    return jsonify(universities)

@app.route('/api/university/diplomas')
def api_get_university_diplomas():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    university_code = session.get('university_code')
    diplomas = get_university_diplomas(university_code)
    return jsonify(diplomas)

@app.route('/api/report_suspicious_activity', methods=['POST'])
def api_report_suspicious_activity():
    data = request.get_json()

    university_code = data.get('university_code')
    diploma_number = data.get('diploma_number', '').strip()
    comment = data.get('comment', '').strip()

    if not university_code or not diploma_number or not comment:
        return jsonify({
            "success": False,
            "message": "Не хватает данных для отправки"
        }), 400

    result = check_diploma_by_params(university_code, diploma_number)

    if not result.get('found'):
        return jsonify({
            "success": False,
            "message": "Диплом не найден"
        }), 404

    if not result.get('valid'):
        return jsonify({
            "success": False,
            "message": "Жалобу можно отправить только для подлинного диплома"
        }), 400

    save_suspicious_report(university_code, diploma_number, comment)

    return jsonify({
        "success": True,
        "message": "Сообщение сохранено"
    })

@app.route('/api/university/add_diploma', methods=['POST'])
def api_add_diploma():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    data = request.get_json()
    university_code = session.get('university_code')
    full_name = data.get('full_name', '').strip()
    graduation_year = data.get('graduation_year', '')
    specialty = data.get('specialty', '').strip()
    diploma_number = data.get('diploma_number', '').strip()
    if not all([full_name, graduation_year, specialty, diploma_number]):
        return jsonify({"success": False, "message": "Заполните все поля"})
    try:
        graduation_year = int(graduation_year)
    except ValueError:
        return jsonify({"success": False, "message": "Год должен быть числом"})

    hash_combined, salt = calculate_diploma_hash(university_code, diploma_number, full_name, graduation_year, specialty)
    diploma_data_for_sign = {
        "university_code": university_code,
        "diploma_number": diploma_number,
        "full_name": full_name,
        "graduation_year": graduation_year,
        "specialty": specialty
    }
    digital_signature = sign_diploma(university_code, diploma_data_for_sign)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO diplomas (university_code, full_name, graduation_year, specialty, diploma_number, status, hash_combined, digital_signature, signature_created_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, CURRENT_TIMESTAMP)
        ''', (university_code, full_name, graduation_year, specialty, diploma_number, hash_combined, digital_signature))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "Диплом добавлен и подписан", "hash": hash_combined})
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"success": False, "message": "Диплом с таким номером уже существует"})


@app.route('/api/university/export_excel', methods=['GET'])
@check_blocked_ip
def api_export_excel():
    """Экспорт дипломов ВУЗа в Excel файл"""
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401

    university_code = session.get('university_code')
    diplomas = get_university_diplomas(university_code)

    if not diplomas:
        return jsonify({"error": "Нет дипломов для экспорта"}), 404

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Дипломы"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1b4e8d", end_color="1b4e8d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["№", "ФИО", "Год выпуска", "Специальность", "Номер диплома", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for idx, diploma in enumerate(diplomas, 1):
        row_num = idx + 1
        status_text = "Активен" if diploma['status'] == 1 else "Аннулирован"

        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=diploma['full_name']).border = border
        ws.cell(row=row_num, column=3, value=diploma['graduation_year']).border = border
        ws.cell(row=row_num, column=4, value=diploma['specialty']).border = border
        ws.cell(row=row_num, column=5, value=diploma['diploma_number']).border = border
        ws.cell(row=row_num, column=6, value=status_text).border = border

        if diploma['status'] == 0:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="fee2e2", end_color="fee2e2",
                                                              fill_type="solid")
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    for col in [1, 3, 6]:
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"diplomas_{university_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/university/toggle_status', methods=['POST'])
def api_toggle_status():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    data = request.get_json()
    diploma_id = data.get('diploma_id')
    current_status = data.get('current_status')
    new_status = 0 if current_status == 1 else 1
    university_code = session.get('university_code')
    success = update_diploma_status(diploma_id, new_status, university_code)
    if success:
        return jsonify({"success": True, "new_status": new_status})
    else:
        return jsonify({"success": False, "message": "Ошибка при обновлении"})

@app.route('/api/university/clear_all_diplomas', methods=['DELETE'])
def api_clear_all_diplomas():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    university_code = session.get('university_code')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM diplomas WHERE university_code = ?", (university_code,))
    deleted_count = cursor.rowcount
    conn.commit()
    conn.close()
    return jsonify({"success": True, "deleted_count": deleted_count})

@app.route('/api/university/delete_diploma', methods=['DELETE'])
def api_delete_diploma():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    data = request.get_json()
    diploma_id = data.get('diploma_id')
    university_code = session.get('university_code')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM diplomas WHERE id = ? AND university_code = ?", (diploma_id, university_code))
    deleted = cursor.rowcount > 0
    conn.commit()
    conn.close()
    if deleted:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "Диплом не найден"})

@app.route('/api/university/upload_excel', methods=['POST'])
@limiter.limit("10 per day", error_message="Слишком много загрузок. Подождите до завтра.")
@check_blocked_ip
def api_upload_excel():
    if 'role' not in session or session.get('role') != 'university':
        return jsonify({"error": "Не авторизован"}), 401
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Файл не выбран"})
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "Файл не выбран"})
    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({"success": False, "message": "Поддерживаются только .xlsx, .xls и .csv"})
    try:
        file_content = file.read()
        if filename.endswith('.csv'):
            encodings = ['utf-8', 'cp1251', 'windows-1251', 'latin1']
            separators = [',', ';', '\t', '|']
            df = None
            for encoding in encodings:
                for sep in separators:
                    try:
                        df = pd.read_csv(io.BytesIO(file_content), header=None, encoding=encoding, sep=sep)
                        if df is not None and len(df.columns) >= 2 and len(df) > 0:
                            break
                    except:
                        continue
                if df is not None and len(df.columns) >= 2:
                    break
            if df is None or len(df.columns) < 2:
                return jsonify({"success": False, "message": "Не удалось прочитать CSV файл"})
        else:
            df = pd.read_excel(io.BytesIO(file_content), header=None)
        if len(df) < 1:
            return jsonify({"success": False, "message": "Файл не содержит данных"})
        university_code = session.get('university_code')
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        added_count = 0
        error_count = 0
        errors = []
        for idx in range(len(df)):
            row = df.iloc[idx]
            if len(row) < 4:
                error_count += 1
                errors.append(f"Строка {idx+1}: недостаточно колонок (нужно 4)")
                continue
            full_name = str(row[0]).strip() if pd.notna(row[0]) else ''
            graduation_year = row[1] if pd.notna(row[1]) else None
            specialty = str(row[2]).strip() if pd.notna(row[2]) else ''
            diploma_number = str(row[3]).strip() if pd.notna(row[3]) else ''
            if not full_name or not diploma_number or graduation_year is None:
                error_count += 1
                errors.append(f"Строка {idx+1}: пропущены обязательные поля")
                continue
            skip_words = ['ФИО', 'FIO', 'ФИО СТУДЕНТА', 'СТУДЕНТ', 'ИМЯ', 'NAME', 'ФАМИЛИЯ']
            if full_name.upper() in [w.upper() for w in skip_words]:
                continue
            try:
                graduation_year = int(graduation_year)
            except (ValueError, TypeError):
                error_count += 1
                errors.append(f"Строка {idx+1}: год должен быть числом")
                continue
            try:
                cursor.execute('''
                    INSERT INTO diplomas (university_code, full_name, graduation_year, specialty, diploma_number, status)
                    VALUES (?, ?, ?, ?, ?, 1)
                ''', (university_code, full_name, graduation_year, specialty, diploma_number))
                added_count += 1
            except sqlite3.IntegrityError:
                error_count += 1
                errors.append(f"Строка {idx+1}: диплом с номером '{diploma_number}' уже существует")
        conn.commit()
        conn.close()
        return jsonify({"success": True, "added_count": added_count, "error_count": error_count, "errors": errors[:10]})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Ошибка при обработке файла: {str(e)}"})

# ---------- API для проверки дипломов и QR ----------
@app.route('/api/search_diploma', methods=['POST'])
@limiter.limit("30 per minute", error_message="Слишком много запросов. Подождите немного.")
@limiter.limit("300 per hour", error_message="Достигнут лимит запросов в час.")
@limiter.limit("1000 per day", error_message="Достигнут дневной лимит запросов.")
@check_blocked_ip
def api_search_diploma():
    data = request.get_json()
    temp_token = data.get('temp_token')

    if temp_token:
        try:
            token_data = serializer.loads(temp_token)
            expiry = token_data.get('expiry')
            if expiry is None:
                return jsonify({"valid": False, "expired": False, "message": "Некорректный токен"})
            if time.time() > expiry:
                return jsonify({"valid": False, "expired": True, "message": "Срок действия ссылки истёк"})
            university_code = token_data.get('university_code')
            diploma_number = token_data.get('diploma_number')
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('SELECT active_token FROM diplomas WHERE university_code = ? AND diploma_number = ?',
                           (university_code, diploma_number))
            row = cursor.fetchone()
            conn.close()
            if row and row[0] != temp_token:
                return jsonify({"valid": False, "expired": True, "message": "QR-код был отозван"})
        except BadSignature:
            return jsonify({"valid": False, "expired": False, "message": "Недействительная ссылка"})
    else:
        university_code = data.get('university_code')
        diploma_number = data.get('diploma_number', '').strip()

    if not university_code or not diploma_number:
        return jsonify({"error": "Не указан код вуза или номер диплома"}), 400

    result = check_diploma_by_params(university_code, diploma_number)

    # === ЛОГИРОВАНИЕ ===
    # Получаем diploma_id
    diploma_id = None
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM diplomas WHERE university_code = ? AND diploma_number = ?",
                   (university_code, diploma_number))
    row = cursor.fetchone()
    if row:
        diploma_id = row[0]
    conn.close()

    # Определяем тип проверки
    verification_type = "temp_link" if temp_token else "manual"

    log_verification(
        diploma_id=diploma_id,
        university_code=university_code,
        diploma_number=diploma_number,
        verification_type=verification_type,
        token=temp_token if temp_token else None,
        result=result.get('message', 'unknown')
    )
    # === КОНЕЦ ЛОГИРОВАНИЯ ===

    # Добавляем информацию о цифровой подписи
    if result.get('found') and result.get('digital_signature'):
        diploma_data = {
            "university_code": university_code,
            "diploma_number": diploma_number,
            "full_name": result['full_name'],
            "graduation_year": result['graduation_year'],
            "specialty": result['specialty']
        }
        signature_valid = verify_diploma_signature(university_code, diploma_data, result['digital_signature'])
        result['signature_valid'] = signature_valid
        result[
            'signature_message'] = "🔐 Цифровая подпись действительна" if signature_valid else "❌ Цифровая подпись НЕДЕЙСТВИТЕЛЬНА"
    elif result.get('found'):
        result['signature_valid'] = False
        result['signature_message'] = "⚠️ Цифровая подпись отсутствует"

    return jsonify(result)


@app.route('/api/scan_qr', methods=['POST'])
@limiter.limit("20 per minute", error_message="Слишком много попыток сканирования.")
@limiter.limit("200 per hour", error_message="Достигнут лимит сканирований в час.")
@check_blocked_ip
def api_scan_qr():
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Файл не выбран"})
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "Файл не выбран"})
    temp_path = os.path.join(tempfile.gettempdir(), file.filename)
    file.save(temp_path)
    try:
        image = cv2.imread(temp_path)
        if image is None:
            return jsonify({"success": False, "message": "Не удалось открыть изображение"})
        detector = cv2.QRCodeDetector()
        data, points, _ = detector.detectAndDecode(image)
        os.remove(temp_path)
        if not data:
            return jsonify({"success": False, "message": "QR-код не найден или не удалось декодировать"})
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(data)
        params = parse_qs(parsed.query)
        token = params.get('token', [None])[0]
        uni_code = params.get('uni_code', [None])[0]
        dip_num = params.get('dip_num', [None])[0]

        if token:
            try:
                token_data = serializer.loads(token)
                expiry = token_data.get('expiry')
                if expiry and time.time() > expiry:
                    return jsonify({"success": False, "message": "Срок действия ссылки истёк"})
                university_code = token_data.get('university_code')
                diploma_number = token_data.get('diploma_number')
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                cursor.execute('SELECT active_token FROM diplomas WHERE university_code = ? AND diploma_number = ?',
                               (university_code, diploma_number))
                row = cursor.fetchone()
                conn.close()
                if row and row[0] != token:
                    return jsonify({"success": False, "message": "QR-код был отозван"})
                if university_code and diploma_number:
                    # === ЛОГИРОВАНИЕ QR ===
                    log_verification(
                        diploma_id=None,
                        university_code=university_code,
                        diploma_number=diploma_number,
                        verification_type="qr_scan",
                        token=token,
                        result="scanned"
                    )
                    # === КОНЕЦ ЛОГИРОВАНИЯ ===
                    return jsonify(
                        {"success": True, "university_code": int(university_code), "diploma_number": diploma_number})
                else:
                    return jsonify({"success": False, "message": "Токен не содержит данных о дипломе"})
            except BadSignature:
                return jsonify({"success": False, "message": "Недействительная ссылка"})
            except Exception as e:
                return jsonify({"success": False, "message": f"Ошибка проверки токена: {str(e)}"})
        elif uni_code and dip_num:
            # === ЛОГИРОВАНИЕ QR (старый формат) ===
            log_verification(
                diploma_id=None,
                university_code=int(uni_code),
                diploma_number=dip_num,
                verification_type="qr_scan",
                token=None,
                result="scanned"
            )
            # === КОНЕЦ ЛОГИРОВАНИЯ ===
            return jsonify({"success": True, "university_code": int(uni_code), "diploma_number": dip_num})
        else:
            return jsonify({"success": False, "message": "QR-код не содержит данных о дипломе"})
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"success": False, "message": f"Ошибка при обработке: {str(e)}"})

@app.route('/api/generate_qr', methods=['POST'])
def api_generate_qr():
    """Генерирует QR-код из URL и возвращает base64 изображение"""
    data = request.get_json()
    url = data.get('url')
    if not url:
        return jsonify({"success": False, "message": "URL не указан"}), 400
    try:
        qr = qrcode.QRCode(version=5, error_correction=qrcode.constants.ERROR_CORRECT_H,
                           box_size=12, border=4)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
        width, height = img.size
        img = img.resize((width * 2, height * 2))
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        return jsonify({"success": True, "qr_image": f"data:image/png;base64,{img_base64}"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/generate_temp_link', methods=['POST'])
@limiter.limit("10 per hour", error_message="Слишком много запросов на создание ссылок.")
@limiter.limit("30 per day", error_message="Достигнут дневной лимит создания ссылок.")
@check_blocked_ip
def api_generate_temp_link():
    data = request.get_json()
    university_code = data.get('university_code')
    diploma_number = data.get('diploma_number', '').strip()
    expiry_seconds = data.get('expiry_seconds', 3600)
    if not university_code or not diploma_number:
        return jsonify({"error": "Не указан код вуза или номер диплома"}), 400
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT full_name, status, graduation_year, specialty FROM diplomas WHERE university_code = ? AND diploma_number = ?',
                   (university_code, diploma_number))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "Диплом не найден"}), 404
    full_name, status, grad_year, specialty = row
    max_seconds = 30 * 24 * 60 * 60
    if expiry_seconds > max_seconds:
        expiry_seconds = max_seconds
    min_seconds = 60
    if expiry_seconds < min_seconds:
        expiry_seconds = min_seconds
    expiry_timestamp = time.time() + expiry_seconds
    token = serializer.dumps({
        "university_code": university_code,
        "diploma_number": diploma_number,
        "full_name": full_name,
        "expiry": expiry_timestamp
    })
    cursor.execute('UPDATE diplomas SET active_token = ? WHERE university_code = ? AND diploma_number = ?',
                   (token, university_code, diploma_number))
    conn.commit()
    conn.close()
    temp_link = f"{request.host_url}?token={token}"
    total_hours = expiry_seconds / 3600
    days = int(total_hours // 24)
    hours = int(total_hours % 24)
    return jsonify({"success": True, "link": temp_link,
                    "expires_in_seconds": expiry_seconds,
                    "expires_in_days": days, "expires_in_hours": hours})

@app.route('/api/revoke_qr', methods=['POST'])
def api_revoke_qr():
    """Отзывает текущий QR-код (удаляет active_token)"""
    data = request.get_json()
    university_code = data.get('university_code')
    diploma_number = data.get('diploma_number', '').strip()
    if not university_code or not diploma_number:
        return jsonify({"success": False, "message": "Не указаны данные"}), 400
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE diplomas SET active_token = NULL WHERE university_code = ? AND diploma_number = ?',
                   (university_code, diploma_number))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "QR-код отозван"})

# ---------- API для HR-порталов ----------
@app.route('/api/verify_diploma_for_hr', methods=['POST'])
@limiter.limit("100 per minute", error_message="Слишком много запросов. Подождите немного.")
@limiter.limit("1000 per hour", error_message="Достигнут лимит запросов в час.")
@limiter.limit("10000 per day", error_message="Достигнут дневной лимит запросов.")
def api_verify_diploma_for_hr():
    data = request.get_json()
    diploma_number = data.get('diploma_number', '').strip()
    university_code = data.get('university_code')
    api_key = data.get('api_key', '').strip()
    full_name_from_request = data.get('full_name', '').strip()

    if not diploma_number:
        return jsonify({"error": "Не указан номер диплома", "status": "error"}), 400
    if not university_code:
        return jsonify({"error": "Не указан код вуза", "status": "error"}), 400

    HR_API_KEYS = os.environ.get('HR_API_KEYS', '').split(',')
    if HR_API_KEYS and HR_API_KEYS[0]:
        if api_key not in HR_API_KEYS:
            log_security_event(request.remote_addr, '/api/verify_diploma_for_hr')
            return jsonify({"error": "Недействительный API ключ", "status": "error"}), 401

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT d.full_name, d.status, d.graduation_year, d.specialty, u.name as university_name
        FROM diplomas d
        JOIN universities u ON d.university_code = u.university_code
        WHERE d.diploma_number = ? AND d.university_code = ?
    ''', (diploma_number, university_code))
    row = cursor.fetchone()
    conn.close()

    response = {
        "diploma_number": diploma_number,
        "university_code": university_code,
        "verified_at": datetime.now().isoformat()
    }

    if not row:
        response.update({"status": "not_found", "message": "Сведений не найдено"})
        return jsonify(response), 404

    full_name_db, status, graduation_year, specialty, university_name = row

    if full_name_from_request:
        normalized_request = ' '.join(full_name_from_request.lower().split())
        normalized_db = ' '.join(full_name_db.lower().split())
        if normalized_request != normalized_db:
            log_security_event(request.remote_addr, '/api/verify_diploma_for_hr',
                               f"ФИО не совпадают: {full_name_from_request} != {full_name_db}")
            response.update({
                "status": "invalid",
                "message": "ФИО владельца диплома не совпадает",
                "full_name_expected": full_name_db
            })
            return jsonify(response), 200

    if status == 1:
        response.update({
            "status": "valid",
            "message": "Подлинный",
            "university_name": university_name,
            "full_name": full_name_db,
            "graduation_year": graduation_year,
            "specialty": specialty
        })
        return jsonify(response), 200
    else:
        response.update({
            "status": "invalid",
            "message": "Аннулирован",
            "university_name": university_name,
            "full_name": full_name_db
        })
        return jsonify(response), 200

@app.route('/api/verify_diplomas_batch', methods=['POST'])
@limiter.limit("10 per minute", error_message="Слишком много запросов. Подождите немного.")
@limiter.limit("100 per hour", error_message="Достигнут лимит запросов в час.")
def api_verify_diplomas_batch():
    data = request.get_json()
    diplomas_list = data.get('diplomas', [])
    api_key = data.get('api_key', '').strip()

    MAX_BATCH_SIZE = 100
    if len(diplomas_list) > MAX_BATCH_SIZE:
        return jsonify({
            "error": f"Слишком много дипломов. Максимум {MAX_BATCH_SIZE} за запрос",
            "status": "error"
        }), 400

    HR_API_KEYS = os.environ.get('HR_API_KEYS', '').split(',')
    if HR_API_KEYS and HR_API_KEYS[0]:
        if api_key not in HR_API_KEYS:
            log_security_event(request.remote_addr, '/api/verify_diplomas_batch')
            return jsonify({"error": "Недействительный API ключ", "status": "error"}), 401

    results = []
    valid_count = 0
    invalid_count = 0
    not_found_count = 0

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    for item in diplomas_list:
        diploma_number = item.get('diploma_number', '').strip()
        university_code = item.get('university_code')

        if not diploma_number or not university_code:
            results.append({
                "diploma_number": diploma_number,
                "university_code": university_code,
                "status": "error",
                "message": "Не указан номер диплома или код вуза"
            })
            continue

        cursor.execute('''
            SELECT status, full_name
            FROM diplomas 
            WHERE diploma_number = ? AND university_code = ?
        ''', (diploma_number, university_code))
        row = cursor.fetchone()

        if not row:
            not_found_count += 1
            results.append({
                "diploma_number": diploma_number,
                "university_code": university_code,
                "status": "not_found",
                "message": "Сведений не найдено"
            })
        elif row[0] == 1:
            valid_count += 1
            results.append({
                "diploma_number": diploma_number,
                "university_code": university_code,
                "status": "valid",
                "message": "Подлинный",
                "full_name": row[1]
            })
        else:
            invalid_count += 1
            results.append({
                "diploma_number": diploma_number,
                "university_code": university_code,
                "status": "invalid",
                "message": "Аннулирован",
                "full_name": row[1]
            })

    conn.close()

    return jsonify({
        "results": results,
        "total": len(diplomas_list),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "not_found_count": not_found_count,
        "verified_at": datetime.now().isoformat()
    })

# ---------- Аутентификация ----------
@app.route('/login/student', methods=['POST'])
@limiter.limit("20 per minute", error_message="Слишком много попыток входа.")
@limiter.limit("100 per hour", error_message="Достигнут лимит попыток входа.")
@check_blocked_ip
def login_student():
    data = request.get_json()
    full_name = data.get('full_name', '').strip()
    diploma_number = data.get('diploma_number', '').strip()
    if check_student_auth(full_name, diploma_number):
        session['role'] = 'student'
        session['full_name'] = full_name
        session['diploma_number'] = diploma_number
        return jsonify({"success": True, "redirect": "/student"})
    else:
        ip = request.remote_addr
        log_security_event(ip, '/login/student')
        return jsonify({"success": False, "message": "Неверное ФИО или номер диплома"})

@app.route('/login/university', methods=['POST'])
@limiter.limit("10 per minute", error_message="Слишком много попыток входа.")
@limiter.limit("30 per hour", error_message="Достигнут лимит попыток входа.")
@check_blocked_ip
def login_university():
    data = request.get_json()
    login = data.get('login', '').strip()
    password = data.get('password', '')
    result = check_university_auth(login, password)
    if result["success"]:
        session['role'] = 'university'
        session['university_code'] = result["university_code"]
        session['university_name'] = result["name"]
        session.modified = True
        return jsonify({"success": True, "redirect": "/university"})
    else:
        ip = request.remote_addr
        log_security_event(ip, '/login/university')
        return jsonify({"success": False, "message": result["message"]})

@app.route('/api/student/status')
def api_student_status():
    if 'role' not in session or session.get('role') != 'student':
        return jsonify({"error": "Не авторизован"}), 401
    full_name = session.get('full_name')
    diploma_number = session.get('diploma_number')
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT status FROM diplomas WHERE full_name = ? AND diploma_number = ?',
                   (full_name, diploma_number))
    row = cursor.fetchone()
    conn.close()
    if row:
        return jsonify({"status": row[0]})
    else:
        return jsonify({"status": None})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('dashboard'))

# ---------- Дебажный эндпоинт ----------
@app.route('/debug/ip')
def debug_ip():
    return jsonify({"ip": request.remote_addr, "headers": dict(request.headers)})

# ---------- Запуск ----------
def main():
    create_security_tables()
    update_db_schema()
    print("\n🚀 Запуск веб-сервера...")
    print("📍 Откройте в браузере: http://127.0.0.1:5000")
    print("🛑 Для остановки нажмите Ctrl+C\n")
    app.run(debug=True, host='0.0.0.0', port=5000)

if __name__ == "__main__":
    main()