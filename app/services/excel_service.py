import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db import get_db_connection
from app.repositories.diploma_repository import get_or_create_student_account, insert_diploma_with_cursor, normalize_student_name
from app.services.auth_service import generate_student_secret, hash_password
from app.services.crypto_service import calculate_diploma_hash, sign_diploma


def safe_excel_value(value):
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def build_excel_export(diplomas: list[dict], university_code):
    wb = Workbook()
    ws = wb.active
    ws.title = "Дипломы"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1b4e8d", end_color="1b4e8d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["№", "ФИО", "Год выпуска", "Специальность", "Номер диплома", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for idx, diploma in enumerate(diplomas, 1):
        row_num = idx + 1
        status_text = "Активен" if diploma["status"] == 1 else "Аннулирован"
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=safe_excel_value(diploma["full_name"])).border = border
        ws.cell(row=row_num, column=3, value=diploma["graduation_year"]).border = border
        ws.cell(row=row_num, column=4, value=safe_excel_value(diploma["specialty"])).border = border
        ws.cell(row=row_num, column=5, value=safe_excel_value(diploma["diploma_number"])).border = border
        ws.cell(row=row_num, column=6, value=status_text).border = border
        if diploma["status"] == 0:
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")

    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    for col in [1, 3, 6]:
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def load_dataframe_from_uploaded_file(file_storage):
    filename = file_storage.filename.lower()
    file_content = file_storage.read()

    if filename.endswith(".csv"):
        encodings = ["utf-8", "cp1251", "windows-1251", "latin1"]
        separators = [",", ";", "	", "|"]
        df = None
        for encoding in encodings:
            for separator in separators:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), header=None, encoding=encoding, sep=separator)
                    if df is not None and len(df.columns) >= 2 and len(df) > 0:
                        break
                except Exception:
                    continue
            if df is not None and len(df.columns) >= 2:
                break
        if df is None or len(df.columns) < 2:
            raise ValueError("Не удалось прочитать CSV файл")
        return df

    return pd.read_excel(io.BytesIO(file_content), header=None)


def bulk_import_diplomas(df, university_code):
    conn = get_db_connection()
    cursor = conn.cursor()
    added_count = 0
    error_count = 0
    reused_account_count = 0
    new_account_count = 0
    errors = []
    credentials = []
    known_new_passwords = {}

    try:
        for idx in range(len(df)):
            row = df.iloc[idx]
            if len(row) < 4:
                error_count += 1
                errors.append(f"Строка {idx + 1}: недостаточно колонок (нужно минимум 4)")
                continue

            full_name = str(row[0]).strip() if pd.notna(row[0]) else ""
            graduation_year = row[1] if pd.notna(row[1]) else None
            specialty = str(row[2]).strip() if pd.notna(row[2]) else ""
            diploma_number = str(row[3]).strip() if pd.notna(row[3]) else ""

            if not full_name or not diploma_number or graduation_year is None:
                error_count += 1
                errors.append(f"Строка {idx + 1}: пропущены обязательные поля")
                continue

            skip_words = ["ФИО", "FIO", "ФИО СТУДЕНТА", "СТУДЕНТ", "ИМЯ", "NAME", "ФАМИЛИЯ"]
            if full_name.upper() in [word.upper() for word in skip_words]:
                continue

            try:
                graduation_year = int(graduation_year)
            except (ValueError, TypeError):
                error_count += 1
                errors.append(f"Строка {idx + 1}: год должен быть числом")
                continue

            try:
                hash_combined, _salt = calculate_diploma_hash(university_code, diploma_number, full_name, graduation_year, specialty)
                diploma_data_for_sign = {
                    "university_code": university_code,
                    "diploma_number": diploma_number,
                    "full_name": full_name,
                    "graduation_year": graduation_year,
                    "specialty": specialty,
                }
                digital_signature = sign_diploma(university_code, diploma_data_for_sign)

                normalized = normalize_student_name(full_name)
                existing_account = get_or_create_student_account(full_name, conn=conn)
                if existing_account:
                    student_account = existing_account
                    reused_account_count += 1
                    student_secret = known_new_passwords.get(normalized)
                else:
                    student_secret = generate_student_secret()
                    student_secret_hash = hash_password(student_secret)
                    student_account = get_or_create_student_account(full_name, secret_hash=student_secret_hash, conn=conn)
                    new_account_count += 1
                    known_new_passwords[normalized] = student_secret

                insert_diploma_with_cursor(
                    cursor=cursor,
                    university_code=university_code,
                    full_name=full_name,
                    graduation_year=graduation_year,
                    specialty=specialty,
                    diploma_number=diploma_number,
                    hash_combined=hash_combined,
                    digital_signature=digital_signature,
                    student_secret_hash=student_account["secret_hash"],
                    student_account_id=student_account["id"],
                )
                credentials.append({
                    "full_name": full_name,
                    "diploma_number": diploma_number,
                    "password": student_secret,
                    "is_new_account": student_secret is not None,
                })
                added_count += 1
            except Exception as error:
                if error.__class__.__name__ == "IntegrityError":
                    error_count += 1
                    errors.append(f"Строка {idx + 1}: диплом с номером '{diploma_number}' уже существует")
                else:
                    error_count += 1
                    errors.append(f"Строка {idx + 1}: {error}")
        conn.commit()
    finally:
        conn.close()

    lines = ["Доступы студентов", ""]
    for item in credentials:
        lines.append(f"ФИО: {item['full_name']}")
        lines.append(f"Логин: {item['diploma_number']}")
        if item['password']:
            lines.append(f"Пароль: {item['password']}")
        else:
            lines.append("Пароль: ранее выданный пароль (не изменялся)")
        lines.append("")

    return {
        "added_count": added_count,
        "error_count": error_count,
        "reused_account_count": reused_account_count,
        "new_account_count": new_account_count,
        "errors": errors[:10],
        "credentials": credentials,
        "password_file_name": "student_access_bulk.txt",
        "password_file_content": "\n".join(lines),
    }
